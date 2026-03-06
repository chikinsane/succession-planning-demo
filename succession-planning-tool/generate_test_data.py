#!/usr/bin/env python3
"""
TalentForge Test Data Generator
Generates 500 synthetic employees for an Indian manufacturing setup.
"""

import random
import math
from datetime import date, timedelta
from collections import defaultdict, Counter
import openpyxl
from openpyxl.utils import get_column_letter

random.seed(42)

# ─── NAME POOLS ───────────────────────────────────────────────────────────────

MALE_FIRST_NAMES = [
    # Hindi/North Indian
    "Rajesh", "Amit", "Vikram", "Sunil", "Manoj", "Arun", "Deepak", "Sanjay",
    "Pradeep", "Ramesh", "Ashok", "Vinod", "Suresh", "Rakesh", "Mukesh",
    "Nitin", "Ajay", "Vijay", "Ravi", "Anil", "Sandeep", "Gaurav", "Rohit",
    "Naveen", "Pankaj", "Yogesh", "Dinesh", "Harish", "Manish", "Sachin",
    "Aarav", "Arjun", "Vivek", "Kunal", "Abhishek", "Rahul",
    # Tamil
    "Karthik", "Senthil", "Murugan", "Balaji", "Saravanan", "Gopal",
    "Venkatesh", "Shankar", "Ramachandran", "Raghav", "Aravind", "Prasad",
    # Telugu
    "Srikanth", "Nagaraj", "Raju", "Chandra", "Krishna", "Mahesh",
    "Srinivas", "Venkat", "Harikrishna", "Jagadish",
    # Bengali
    "Subhash", "Soumya", "Debashish", "Partha", "Aniket", "Sourav",
    "Dipankar", "Arnab", "Biswajit", "Tarun",
    # Marathi
    "Ganesh", "Sagar", "Tushar", "Omkar", "Nikhil", "Pratik",
    "Ameya", "Swapnil", "Tejas", "Vikas",
    # Gujarati
    "Hitesh", "Jignesh", "Chirag", "Darshan", "Ketan", "Jigar",
    "Mihir", "Neel", "Pranav", "Yash",
]

FEMALE_FIRST_NAMES = [
    # Hindi/North Indian
    "Priya", "Anjali", "Sunita", "Rekha", "Kavita", "Neha", "Pooja",
    "Shweta", "Divya", "Swati", "Ritu", "Nisha", "Deepa", "Pallavi",
    "Shruti", "Meena", "Anita", "Seema", "Jyoti", "Shalini",
    "Aisha", "Tanvi", "Ishita", "Aditi", "Sneha",
    # Tamil
    "Lakshmi", "Bharathi", "Revathi", "Sowmya", "Preethi",
    "Nandhini", "Vaishnavi", "Devi", "Kamala",
    # Telugu
    "Swapna", "Lavanya", "Padma", "Sravani", "Keerthi",
    # Bengali
    "Moumita", "Shreya", "Rina", "Mitali", "Gargi",
    # Marathi
    "Madhavi", "Vaishali", "Sarita", "Ashwini", "Manasi",
    # Gujarati
    "Hetal", "Komal", "Janki", "Bhumika", "Riddhi",
]

LAST_NAMES = [
    "Sharma", "Verma", "Singh", "Patel", "Gupta", "Kumar", "Reddy",
    "Rao", "Nair", "Menon", "Iyer", "Pillai", "Das", "Roy",
    "Banerjee", "Chatterjee", "Mukherjee", "Sen", "Ghosh", "Bose",
    "Desai", "Joshi", "Kulkarni", "Patil", "Deshpande",
    "Shah", "Mehta", "Trivedi", "Bhatt", "Parikh",
    "Mishra", "Pandey", "Dubey", "Tiwari", "Srivastava",
    "Choudhary", "Agarwal", "Saxena", "Kapoor", "Malhotra",
    "Rajan", "Subramaniam", "Naidu", "Prasad", "Hegde",
    "Thakur", "Chauhan", "Yadav", "Rajput", "Mohan",
]

# ─── ORG STRUCTURE ────────────────────────────────────────────────────────────

DEPARTMENTS = [
    "Production", "Quality", "Maintenance", "EHS", "Supply Chain",
    "Engineering", "Automation", "HR", "Finance", "IT"
]

PLANTS = [
    "Plant A — Pune",
    "Plant B — Chennai",
    "Plant C — Ahmedabad",
    "Plant D — Jamshedpur",
    "Plant E — Hosur",
    "Corporate HQ — Mumbai",
]

ROLE_HIERARCHY = {
    "Production": [
        ("VP Operations", "L8"), ("Plant Head", "L7"), ("Production Manager", "L6"),
        ("Shift Supervisor", "L4"), ("Line Leader", "L3"),
        ("Production Engineer", "L4"), ("Senior Operator", "L2"), ("Operator", "L1"),
    ],
    "Quality": [
        ("Quality Head", "L7"), ("Quality Manager", "L6"), ("Quality Engineer", "L4"),
        ("QC Inspector", "L3"), ("Lab Technician", "L2"),
    ],
    "Maintenance": [
        ("Maintenance Head", "L7"), ("Maintenance Manager", "L6"),
        ("Maintenance Engineer", "L4"), ("Technician", "L2"), ("Electrician", "L1"),
    ],
    "EHS": [
        ("EHS Director", "L7"), ("EHS Manager", "L6"), ("Safety Officer", "L4"),
        ("Safety Coordinator", "L3"), ("Environmental Engineer", "L4"),
    ],
    "Supply Chain": [
        ("Supply Chain Head", "L7"), ("Procurement Head", "L6"),
        ("Logistics Manager", "L6"), ("Warehouse Manager", "L5"),
        ("Planner", "L3"), ("Buyer", "L3"),
    ],
    "Engineering": [
        ("Engineering Manager", "L6"), ("Design Engineer", "L4"),
        ("Process Engineer", "L4"), ("Project Engineer", "L4"), ("CAD Technician", "L2"),
    ],
    "Automation": [
        ("Automation Lead", "L6"), ("Automation Engineer", "L4"),
        ("PLC Programmer", "L3"), ("SCADA Engineer", "L4"),
        ("Instrumentation Tech", "L2"),
    ],
    "HR": [
        ("CHRO", "L8"), ("HR Director", "L7"), ("HR Manager", "L6"),
        ("HR Business Partner", "L5"), ("Talent Manager", "L5"), ("Recruiter", "L3"),
    ],
    "Finance": [
        ("CFO", "L8"), ("Finance Controller", "L7"), ("Accounts Manager", "L6"),
        ("Cost Analyst", "L4"), ("Accountant", "L3"),
    ],
    "IT": [
        ("IT Manager", "L6"), ("Systems Administrator", "L4"),
        ("Developer", "L4"), ("Network Engineer", "L4"), ("IT Support", "L2"),
    ],
}

FEMALE_HEAVY_DEPTS = {"Quality", "HR", "Finance", "IT"}

GRADE_NUM = {"L1": 1, "L2": 2, "L3": 3, "L4": 4, "L5": 5, "L6": 6, "L7": 7, "L8": 8}

EDUCATION_BY_GRADE = {
    "L1": ["ITI", "Diploma", "ITI", "ITI"],
    "L2": ["ITI", "Diploma", "B.Sc", "Diploma"],
    "L3": ["Diploma", "B.Sc", "B.Com", "B.Tech"],
    "L4": ["B.Tech", "B.Sc", "B.Com", "Diploma", "B.Tech"],
    "L5": ["B.Tech", "MBA", "M.Tech", "B.Com", "M.Com"],
    "L6": ["B.Tech", "MBA", "M.Tech", "MBA"],
    "L7": ["MBA", "M.Tech", "B.Tech", "PhD", "MBA"],
    "L8": ["MBA", "M.Tech", "PhD", "MBA"],
}

MOBILITY_OPTIONS = [
    "Open to relocation", "Prefer current location",
    "International ready", "Within state only"
]

ASPIRATION_OPTIONS = [
    "VP Operations", "Plant Head", "Functional Head", "Technical Expert",
    "Same role", "Cross-functional move", "General Management", "CHRO", "CFO"
]

READINESS_OPTIONS = ["Ready Now", "1-2 Years", "3+ Years", "Not Assessed"]
READINESS_WEIGHTS = [0.15, 0.30, 0.25, 0.30]

DEV_STATUS_OPTIONS = ["On Track", "Ahead", "Behind", "Not Started"]

FLIGHT_FACTORS_POOL = [
    "Low pay relativity", "Short tenure", "Stalled career",
    "Market demand", "Poor manager relationship", "Limited growth",
    "Competitor poaching", "Work-life balance"
]

COMPETENCY_POOL = {
    "Production": ["Lean Manufacturing", "Six Sigma", "Production Planning", "Shop Floor Management",
                    "Capacity Planning", "OEE Optimization", "Kaizen", "TPM"],
    "Quality": ["Statistical Process Control", "Quality Auditing", "ISO Standards", "Root Cause Analysis",
                 "FMEA", "Calibration", "APQP", "Metrology"],
    "Maintenance": ["Preventive Maintenance", "Predictive Analytics", "Equipment Reliability",
                     "CMMS Management", "Hydraulics", "Pneumatics", "Vibration Analysis", "Thermography"],
    "EHS": ["OSHA Compliance", "Risk Assessment", "Incident Investigation", "Environmental Management",
             "Fire Safety", "Ergonomics", "Hazard Identification", "Emergency Response"],
    "Supply Chain": ["Inventory Management", "Vendor Management", "Logistics Optimization",
                      "Demand Forecasting", "SAP MM", "Procurement Strategy", "Warehouse Design", "S&OP"],
    "Engineering": ["CAD/CAM", "Product Design", "Process Optimization", "Value Engineering",
                     "GD&T", "FEA Analysis", "DFM/DFA", "Project Management"],
    "Automation": ["PLC Programming", "SCADA Systems", "Industrial IoT", "Robotics",
                    "HMI Design", "DCS Systems", "Instrumentation", "Control Systems"],
    "HR": ["Talent Management", "Organizational Development", "Labor Relations", "Compensation Strategy",
            "Succession Planning", "Employee Engagement", "Learning & Development", "HR Analytics"],
    "Finance": ["Financial Planning", "Cost Accounting", "Budgeting", "Tax Compliance",
                 "Financial Reporting", "Internal Controls", "Working Capital Management", "ERP Systems"],
    "IT": ["Cloud Infrastructure", "Cybersecurity", "ERP Administration", "Network Management",
            "Software Development", "Database Management", "IT Service Management", "Data Analytics"],
}

DEPT_WEIGHTS = {
    "Production": 0.30,
    "Quality": 0.10,
    "Maintenance": 0.10,
    "EHS": 0.06,
    "Supply Chain": 0.08,
    "Engineering": 0.08,
    "Automation": 0.06,
    "HR": 0.08,
    "Finance": 0.07,
    "IT": 0.07,
}

PLANT_WEIGHTS_MANUFACTURING = {
    "Plant A — Pune": 0.25,
    "Plant B — Chennai": 0.22,
    "Plant C — Ahmedabad": 0.20,
    "Plant D — Jamshedpur": 0.18,
    "Plant E — Hosur": 0.15,
}

HQ_DEPARTMENTS = {"HR", "Finance", "IT"}

# ─── HELPER FUNCTIONS ─────────────────────────────────────────────────────────

def clamp(val, lo, hi):
    return max(lo, min(hi, val))

def gen_perf_scores():
    base = clamp(random.gauss(3.4, 0.6), 1.0, 5.0)
    trajectory = random.choice(["stable", "improving", "declining", "stable", "stable"])
    scores = []
    for yr in range(3):
        if trajectory == "improving":
            s = base + yr * random.uniform(0.1, 0.3)
        elif trajectory == "declining":
            s = base - yr * random.uniform(0.1, 0.25)
        else:
            s = base + random.uniform(-0.2, 0.2)
        scores.append(round(clamp(s, 1.0, 5.0), 1))
    return scores

def perf_level(score):
    if score >= 3.8:
        return "High"
    elif score >= 2.5:
        return "Medium"
    return "Low"

def gen_potential(perf_avg):
    pot = perf_avg * 0.6 + random.gauss(1.5, 0.5)
    return round(clamp(pot, 1.0, 5.0), 1)

def gen_flight_risk(compa_ratio, tenure_years, grade_years_stuck, perf_avg, dept):
    """Generate 30/60/90 day flight risk. Tuned for ~15% having r30 > 0.7."""
    base_risk = 0.25  # raised baseline

    # Low pay -> higher risk
    if compa_ratio < 0.85:
        base_risk += 0.35
    elif compa_ratio < 0.92:
        base_risk += 0.18

    # Short tenure (1-3 years) -> higher risk
    if 1 <= tenure_years <= 3:
        base_risk += 0.22

    # Stalled career
    if grade_years_stuck >= 5:
        base_risk += 0.20

    # High performers with low pay have extra risk
    if perf_avg >= 3.8 and compa_ratio < 0.90:
        base_risk += 0.15

    # Market demand for automation/IT
    if dept in ("Automation", "IT"):
        base_risk += 0.15

    # Add noise
    base_risk += random.uniform(-0.12, 0.12)

    r30 = round(clamp(base_risk, 0.0, 1.0), 2)
    r60 = round(clamp(r30 + random.uniform(0.03, 0.12), 0.0, 1.0), 2)
    r90 = round(clamp(r60 + random.uniform(0.03, 0.12), 0.0, 1.0), 2)
    return r30, r60, r90

def gen_flight_factors(r30, compa_ratio, tenure_years, grade_years_stuck):
    factors = []
    if compa_ratio < 0.90:
        factors.append("Low pay relativity")
    if tenure_years <= 3:
        factors.append("Short tenure")
    if grade_years_stuck >= 5:
        factors.append("Stalled career")
    if r30 > 0.4:
        extras = [f for f in FLIGHT_FACTORS_POOL if f not in factors]
        random.shuffle(extras)
        factors.extend(extras[:random.randint(0, 2)])
    return ", ".join(factors[:3])

def gen_knowledge_risk(tenure_years, grade_num):
    base = (tenure_years / 25.0) * 0.4 + (grade_num / 8.0) * 0.4
    base += random.uniform(-0.1, 0.1)
    return round(clamp(base, 0.0, 1.0), 2)

# ─── EMPLOYEE GENERATION ─────────────────────────────────────────────────────

TODAY = date(2025, 3, 6)
HIRE_START = date(2000, 1, 1)
HIRE_END = date(2024, 6, 1)
HIRE_RANGE_DAYS = (HIRE_END - HIRE_START).days

def generate_employees(n=500):
    employees = []
    emp_id_counter = 1001

    dept_counts = {}
    remaining = n
    dept_list = list(DEPT_WEIGHTS.keys())
    for i, dept in enumerate(dept_list):
        if i == len(dept_list) - 1:
            dept_counts[dept] = remaining
        else:
            cnt = round(n * DEPT_WEIGHTS[dept])
            dept_counts[dept] = cnt
            remaining -= cnt

    all_assignments = []

    for dept, count in dept_counts.items():
        roles_in_dept = ROLE_HIERARCHY[dept]
        role_grade_nums = [GRADE_NUM[g] for _, g in roles_in_dept]
        role_weights = [max(1, 9 - g) ** 1.5 for g in role_grade_nums]
        total_w = sum(role_weights)
        role_weights = [w / total_w for w in role_weights]

        for _ in range(count):
            role_title, grade = random.choices(roles_in_dept, weights=role_weights, k=1)[0]
            gnum = GRADE_NUM[grade]

            if dept in HQ_DEPARTMENTS:
                if gnum >= 7 or random.random() < 0.40:
                    plant = "Corporate HQ — Mumbai"
                    bu = "Corporate"
                else:
                    plant = random.choices(
                        list(PLANT_WEIGHTS_MANUFACTURING.keys()),
                        weights=list(PLANT_WEIGHTS_MANUFACTURING.values()), k=1
                    )[0]
                    bu = "Manufacturing"
            else:
                if gnum >= 8 or (gnum >= 7 and random.random() < 0.3):
                    plant = "Corporate HQ — Mumbai"
                    bu = "Corporate"
                else:
                    plant = random.choices(
                        list(PLANT_WEIGHTS_MANUFACTURING.keys()),
                        weights=list(PLANT_WEIGHTS_MANUFACTURING.values()), k=1
                    )[0]
                    bu = "Manufacturing"

            all_assignments.append((dept, plant, role_title, grade, bu))

    random.shuffle(all_assignments)

    for idx, (dept, plant, role_title, grade, bu) in enumerate(all_assignments):
        eid = f"EMP{emp_id_counter}"
        emp_id_counter += 1
        gnum = GRADE_NUM[grade]

        # Gender
        if dept in FEMALE_HEAVY_DEPTS:
            gender_r = random.random()
            if gender_r < 0.45:
                gender = "Female"
            elif gender_r < 0.95:
                gender = "Male"
            else:
                gender = "Non-Binary"
        else:
            gender_r = random.random()
            if gender_r < 0.20:
                gender = "Female"
            elif gender_r < 0.95:
                gender = "Male"
            else:
                gender = "Non-Binary"
        if gnum >= 7 and dept in ("Production", "Maintenance") and gender == "Female":
            if random.random() < 0.7:
                gender = "Male"

        # Name
        if gender == "Female":
            first_name = random.choice(FEMALE_FIRST_NAMES)
        elif gender == "Non-Binary":
            first_name = random.choice(MALE_FIRST_NAMES + FEMALE_FIRST_NAMES)
        else:
            first_name = random.choice(MALE_FIRST_NAMES)
        last_name = random.choice(LAST_NAMES)

        # Hire Date
        if gnum >= 7:
            hire_date = HIRE_START + timedelta(days=random.randint(0, int(HIRE_RANGE_DAYS * 0.5)))
        elif gnum >= 5:
            hire_date = HIRE_START + timedelta(days=random.randint(0, int(HIRE_RANGE_DAYS * 0.75)))
        else:
            hire_date = HIRE_START + timedelta(days=random.randint(0, HIRE_RANGE_DAYS))

        # Jamshedpur aging workforce
        if plant == "Plant D — Jamshedpur" and random.random() < 0.5:
            hire_date = HIRE_START + timedelta(days=random.randint(0, int(HIRE_RANGE_DAYS * 0.4)))

        tenure_days = (TODAY - hire_date).days
        tenure_years = round(tenure_days / 365.25, 1)

        # Age
        min_age_for_grade = {1: 22, 2: 23, 3: 25, 4: 26, 5: 30, 6: 33, 7: 40, 8: 45}
        min_age = max(min_age_for_grade[gnum], 22 + int(tenure_years))
        age = min(62, min_age + random.randint(0, 8))
        if plant == "Plant D — Jamshedpur":
            age = min(62, age + random.randint(0, 5))

        # Education
        education = random.choice(EDUCATION_BY_GRADE[grade])

        # Performance
        perf_scores = gen_perf_scores()
        perf_avg = round(sum(perf_scores) / 3.0, 1)
        perf_lev = perf_level(perf_avg)

        # Potential
        pot_score = gen_potential(perf_avg)
        pot_lev = perf_level(pot_score)

        # Competency scores
        learning_agility = round(clamp(pot_score * 0.5 + random.gauss(1.8, 0.5), 1.0, 5.0), 1)
        strategic_thinking = round(clamp(pot_score * 0.4 + random.gauss(1.7, 0.6), 1.0, 5.0), 1)
        people_leadership = round(clamp(pot_score * 0.45 + random.gauss(1.6, 0.5), 1.0, 5.0), 1)

        # Compa Ratio — deliberately create more underpaid employees
        compa_ratio = round(clamp(random.gauss(0.98, 0.12), 0.70, 1.35), 2)
        if dept in ("Automation", "IT") and random.random() < 0.40:
            compa_ratio = round(clamp(random.gauss(0.82, 0.07), 0.70, 0.95), 2)
        if gnum <= 2 and random.random() < 0.35:
            compa_ratio = round(clamp(random.gauss(0.83, 0.08), 0.70, 0.95), 2)

        # Grade years stuck
        grade_years_stuck = max(0, tenure_years - random.uniform(0, tenure_years * 0.6))

        # Flight Risk
        r30, r60, r90 = gen_flight_risk(compa_ratio, tenure_years, grade_years_stuck, perf_avg, dept)

        flight_factors = gen_flight_factors(r30, compa_ratio, tenure_years, grade_years_stuck)
        if dept == "Automation" and "Market demand" not in flight_factors:
            if flight_factors:
                flight_factors += ", Market demand"
            else:
                flight_factors = "Market demand"

        # Mobility
        if gnum >= 6:
            mobility = random.choices(MOBILITY_OPTIONS, weights=[0.35, 0.25, 0.25, 0.15], k=1)[0]
        else:
            mobility = random.choices(MOBILITY_OPTIONS, weights=[0.20, 0.45, 0.10, 0.25], k=1)[0]

        # Aspiration
        if gnum >= 7:
            aspiration = random.choice(["VP Operations", "General Management", "Functional Head", "Same role"])
        elif gnum >= 5:
            aspiration = random.choice(["Plant Head", "Functional Head", "General Management", "Cross-functional move", "Technical Expert"])
        else:
            aspiration = random.choice(ASPIRATION_OPTIONS)
        if dept == "HR" and gnum >= 6:
            aspiration = random.choice(["CHRO", "Functional Head", "General Management"])
        if dept == "Finance" and gnum >= 6:
            aspiration = random.choice(["CFO", "Functional Head", "General Management"])

        # 360 Score
        score_360 = round(clamp(perf_avg * 0.5 + pot_score * 0.3 + random.gauss(0.8, 0.3), 2.0, 5.0), 1)

        # Readiness
        readiness = random.choices(READINESS_OPTIONS, weights=READINESS_WEIGHTS, k=1)[0]
        if gnum >= 5 and perf_avg >= 3.8 and pot_score >= 3.8:
            readiness = random.choices(["Ready Now", "1-2 Years", "3+ Years"], weights=[0.5, 0.35, 0.15], k=1)[0]

        # Successor For
        successor_for = ""
        if random.random() < 0.40:
            possible_roles = []
            for r, g in ROLE_HIERARCHY[dept]:
                if GRADE_NUM[g] > gnum:
                    possible_roles.append(r)
            if possible_roles:
                successor_for = random.choice(possible_roles)

        # Dev Status
        dev_status = random.choices(DEV_STATUS_OPTIONS, weights=[0.40, 0.15, 0.25, 0.20], k=1)[0]

        # Mentor
        mentor = "Yes" if random.random() < 0.40 else "No"

        # Knowledge Risk
        knowledge_risk = gen_knowledge_risk(tenure_years, gnum)

        employees.append({
            "Employee ID": eid,
            "First Name": first_name,
            "Last Name": last_name,
            "Department": dept,
            "Plant": plant,
            "Business Unit": bu,
            "Role": role_title,
            "Grade": grade,
            "Hire Date": hire_date.strftime("%Y-%m-%d"),
            "Tenure": tenure_years,
            "Age": age,
            "Gender": gender,
            "Education": education,
            "Perf Y1": perf_scores[0],
            "Perf Y2": perf_scores[1],
            "Perf Y3": perf_scores[2],
            "Performance Score": perf_avg,
            "Performance Level": perf_lev,
            "Potential Score": pot_score,
            "Potential Level": pot_lev,
            "Learning Agility": learning_agility,
            "Strategic Thinking": strategic_thinking,
            "People Leadership": people_leadership,
            "Flight Risk 30": r30,
            "Flight Risk 60": r60,
            "Flight Risk 90": r90,
            "Flight Factors": flight_factors,
            "Compa Ratio": compa_ratio,
            "Mobility": mobility,
            "Aspiration": aspiration,
            "360 Score": score_360,
            "Readiness": readiness,
            "Successor For": successor_for,
            "Dev Status": dev_status,
            "Mentor": mentor,
            "Knowledge Risk": knowledge_risk,
        })

    return employees


# ─── ROLES SHEET ──────────────────────────────────────────────────────────────

def generate_roles():
    """Generate ~60 critical role entries spread across plants."""
    roles = []
    role_id = 1

    # For Tier 1/2 roles, create entries across multiple plants
    for dept in DEPARTMENTS:
        hierarchy = ROLE_HIERARCHY[dept]
        for role_title, grade in hierarchy:
            gnum = GRADE_NUM[grade]
            if gnum >= 7:
                tier = 1
            elif gnum >= 5:
                tier = 2
            else:
                tier = 3

            # Decide which plants get this role
            if dept in HQ_DEPARTMENTS:
                if gnum >= 7:
                    plants_for_role = ["Corporate HQ — Mumbai"]
                else:
                    # HQ + 1-2 major plants
                    plants_for_role = ["Corporate HQ — Mumbai"]
                    if tier <= 2:
                        plants_for_role.append("Plant A — Pune")
            else:
                if gnum >= 8:
                    plants_for_role = ["Corporate HQ — Mumbai"]
                elif gnum >= 7:
                    # Directors at HQ or major plants
                    plants_for_role = ["Corporate HQ — Mumbai", "Plant A — Pune"]
                elif tier <= 2:
                    # Managers at each plant
                    plants_for_role = list(PLANT_WEIGHTS_MANUFACTURING.keys())[:3]
                else:
                    plants_for_role = [random.choice(list(PLANT_WEIGHTS_MANUFACTURING.keys()))]

            for plant in plants_for_role:
                bu = "Corporate" if plant == "Corporate HQ — Mumbai" else "Manufacturing"
                critical = "Yes" if tier <= 2 else ("Yes" if random.random() < 0.3 else "No")
                comps = random.sample(COMPETENCY_POOL[dept], min(len(COMPETENCY_POOL[dept]), random.randint(4, 6)))
                competencies_str = ", ".join(comps)

                roles.append({
                    "Role ID": f"R{role_id:03d}",
                    "Title": role_title,
                    "Tier": tier,
                    "Department": dept,
                    "Plant": plant,
                    "Business Unit": bu,
                    "Critical": critical,
                    "Competencies": competencies_str,
                })
                role_id += 1

                # Limit total to ~60
                if role_id > 60:
                    break
            if role_id > 60:
                break
        if role_id > 60:
            break

    # If we haven't reached 60, fill with remaining roles
    if role_id <= 60:
        for dept in DEPARTMENTS:
            hierarchy = ROLE_HIERARCHY[dept]
            for role_title, grade in hierarchy:
                gnum = GRADE_NUM[grade]
                if gnum >= 7:
                    tier = 1
                elif gnum >= 5:
                    tier = 2
                else:
                    tier = 3

                # Check if we already have this role title
                existing_titles = {r["Title"] for r in roles}
                if role_title in existing_titles:
                    continue

                plant = random.choice(list(PLANT_WEIGHTS_MANUFACTURING.keys()))
                bu = "Manufacturing"
                critical = "Yes" if tier <= 2 else "No"
                comps = random.sample(COMPETENCY_POOL[dept], min(len(COMPETENCY_POOL[dept]), random.randint(4, 6)))

                roles.append({
                    "Role ID": f"R{role_id:03d}",
                    "Title": role_title,
                    "Tier": tier,
                    "Department": dept,
                    "Plant": plant,
                    "Business Unit": bu,
                    "Critical": critical,
                    "Competencies": ", ".join(comps),
                })
                role_id += 1
                if role_id > 65:
                    break
            if role_id > 65:
                break

    return roles


# ─── SUCCESSION SHEET ─────────────────────────────────────────────────────────

def generate_succession(employees, roles):
    """Generate succession pipeline — 2-4 successors per critical/tier1-2 role. Target 120-150 rows."""
    succession = []

    dept_employees = defaultdict(list)
    for emp in employees:
        dept_employees[emp["Department"]].append(emp)

    # Use all roles with Critical=Yes or Tier <= 2
    target_roles = [r for r in roles if r["Critical"] == "Yes" or r["Tier"] <= 2]

    # If not enough, also add some Tier 3 roles
    if len(target_roles) < 40:
        tier3 = [r for r in roles if r["Tier"] == 3 and r not in target_roles]
        random.shuffle(tier3)
        target_roles.extend(tier3[:max(0, 40 - len(target_roles))])

    for role in target_roles:
        dept = role["Department"]
        candidates = dept_employees.get(dept, [])
        if not candidates:
            continue

        num_successors = random.randint(2, 4)
        sorted_candidates = sorted(candidates, key=lambda e: e["Performance Score"], reverse=True)

        role_grade = None
        for rt, rg in ROLE_HIERARCHY[dept]:
            if rt == role["Title"]:
                role_grade = GRADE_NUM[rg]
                break
        if role_grade is None:
            role_grade = 5

        eligible = [c for c in sorted_candidates if GRADE_NUM[c["Grade"]] < role_grade]
        if len(eligible) < 2:
            # Include same-grade candidates as fallback
            eligible = [c for c in sorted_candidates if GRADE_NUM[c["Grade"]] <= role_grade]
        if not eligible:
            eligible = sorted_candidates[:6]

        # Pick different employees for each role if possible
        random.shuffle(eligible)
        chosen = eligible[:num_successors]

        for emp in chosen:
            match_score = round(clamp(
                emp["Performance Score"] / 5.0 * 0.4 +
                emp["Potential Score"] / 5.0 * 0.3 +
                random.uniform(0.1, 0.3),
                0.50, 0.98
            ), 2)

            succession.append({
                "Role ID": role["Role ID"],
                "Role": role["Title"],
                "Tier": role["Tier"],
                "Plant": role["Plant"],
                "Department": dept,
                "Employee ID": emp["Employee ID"],
                "Employee Name": f"{emp['First Name']} {emp['Last Name']}",
                "Readiness": emp["Readiness"],
                "Match Score": match_score,
            })

    return succession


# ─── ORG STRUCTURE SHEET ──────────────────────────────────────────────────────

def generate_org_structure(employees):
    combos = set()
    for emp in employees:
        combos.add((emp["Department"], emp["Plant"], emp["Business Unit"]))
    return sorted([{"Department": d, "Plant": p, "Business Unit": b} for d, p, b in combos],
                  key=lambda x: (x["Department"], x["Plant"]))


# ─── WRITE EXCEL ──────────────────────────────────────────────────────────────

def write_excel(employees, roles, succession, org_structure, filepath):
    wb = openpyxl.Workbook()

    # Sheet 1: Employees
    ws1 = wb.active
    ws1.title = "Employees"
    emp_cols = [
        "Employee ID", "First Name", "Last Name", "Department", "Plant",
        "Business Unit", "Role", "Grade", "Hire Date", "Tenure", "Age",
        "Gender", "Education", "Perf Y1", "Perf Y2", "Perf Y3",
        "Performance Score", "Performance Level", "Potential Score",
        "Potential Level", "Learning Agility", "Strategic Thinking",
        "People Leadership", "Flight Risk 30", "Flight Risk 60",
        "Flight Risk 90", "Flight Factors", "Compa Ratio", "Mobility",
        "Aspiration", "360 Score", "Readiness", "Successor For",
        "Dev Status", "Mentor", "Knowledge Risk"
    ]
    ws1.append(emp_cols)
    for emp in employees:
        ws1.append([emp[c] for c in emp_cols])

    # Sheet 2: Roles
    ws2 = wb.create_sheet("Roles")
    role_cols = ["Role ID", "Title", "Tier", "Department", "Plant", "Business Unit", "Critical", "Competencies"]
    ws2.append(role_cols)
    for role in roles:
        ws2.append([role[c] for c in role_cols])

    # Sheet 3: Succession
    ws3 = wb.create_sheet("Succession")
    succ_cols = ["Role ID", "Role", "Tier", "Plant", "Department", "Employee ID", "Employee Name", "Readiness", "Match Score"]
    ws3.append(succ_cols)
    for s in succession:
        ws3.append([s[c] for c in succ_cols])

    # Sheet 4: Org Structure
    ws4 = wb.create_sheet("Org Structure")
    org_cols = ["Department", "Plant", "Business Unit"]
    ws4.append(org_cols)
    for o in org_structure:
        ws4.append([o[c] for c in org_cols])

    # Auto-width columns
    for ws in [ws1, ws2, ws3, ws4]:
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    wb.save(filepath)
    print(f"File saved: {filepath}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Generating employees...")
    employees = generate_employees(500)
    print(f"  Employees generated: {len(employees)}")

    print("Generating roles...")
    roles = generate_roles()
    print(f"  Roles generated: {len(roles)}")

    print("Generating succession pipeline...")
    succession = generate_succession(employees, roles)
    print(f"  Succession entries: {len(succession)}")

    print("Generating org structure...")
    org_structure = generate_org_structure(employees)
    print(f"  Org structure entries: {len(org_structure)}")

    filepath = "TalentForge_TestData_500.xlsx"
    write_excel(employees, roles, succession, org_structure, filepath)

    # Print summary stats
    print("\n" + "="*60)
    print("FINAL SUMMARY")
    print("="*60)
    print(f"Sheet 'Employees': {len(employees)} rows")
    print(f"Sheet 'Roles': {len(roles)} rows")
    print(f"Sheet 'Succession': {len(succession)} rows")
    print(f"Sheet 'Org Structure': {len(org_structure)} rows")

    dept_dist = Counter(e["Department"] for e in employees)
    gender_dist = Counter(e["Gender"] for e in employees)
    plant_dist = Counter(e["Plant"] for e in employees)
    grade_dist = Counter(e["Grade"] for e in employees)
    readiness_dist = Counter(e["Readiness"] for e in employees)
    high_fr30 = sum(1 for e in employees if e["Flight Risk 30"] > 0.7)

    print(f"\nDepartment distribution:")
    for d in sorted(dept_dist.keys()):
        print(f"  {d}: {dept_dist[d]}")

    print(f"\nGender distribution:")
    for g in sorted(gender_dist.keys()):
        pct = gender_dist[g] / len(employees) * 100
        print(f"  {g}: {gender_dist[g]} ({pct:.1f}%)")

    print(f"\nPlant distribution:")
    for p in sorted(plant_dist.keys()):
        print(f"  {p}: {plant_dist[p]}")

    print(f"\nGrade distribution:")
    for g in sorted(grade_dist.keys()):
        print(f"  {g}: {grade_dist[g]}")

    print(f"\nReadiness distribution:")
    for r in sorted(readiness_dist.keys()):
        pct = readiness_dist[r] / len(employees) * 100
        print(f"  {r}: {readiness_dist[r]} ({pct:.1f}%)")

    print(f"\nFlight Risk 30 > 0.7: {high_fr30} ({high_fr30/len(employees)*100:.1f}%)")

    avg_age = sum(e["Age"] for e in employees) / len(employees)
    avg_tenure = sum(e["Tenure"] for e in employees) / len(employees)
    print(f"Average age: {avg_age:.1f}")
    print(f"Average tenure: {avg_tenure:.1f} years")

    # Succession coverage story
    print(f"\nSuccession pipeline Tier breakdown:")
    tier_dist = Counter(s["Tier"] for s in succession)
    for t in sorted(tier_dist.keys()):
        print(f"  Tier {t}: {tier_dist[t]} entries")

    # Plant succession coverage
    print(f"\nSuccession entries by plant:")
    plant_succ = Counter(s["Plant"] for s in succession)
    for p in sorted(plant_succ.keys()):
        print(f"  {p}: {plant_succ[p]}")

    # Gender in senior roles (L6+)
    senior = [e for e in employees if GRADE_NUM[e["Grade"]] >= 6]
    senior_gender = Counter(e["Gender"] for e in senior)
    print(f"\nGender in senior roles (L6+): {dict(senior_gender)}")

    # Dept flight risk
    print(f"\nAvg Flight Risk 30 by department:")
    for dept in sorted(DEPARTMENTS):
        dept_emps = [e for e in employees if e["Department"] == dept]
        avg_fr = sum(e["Flight Risk 30"] for e in dept_emps) / len(dept_emps)
        print(f"  {dept}: {avg_fr:.2f}")
